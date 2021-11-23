#forras: https://www.marsja.se/your-guide-to-reading-excel-xlsx-files-in-python/

import sys
import openpyxl
file_to_read = sys.argv[1]
file_content = openpyxl.load_workbook(file_to_read)
sheet = file_content.active

#egy ertek kinyomtatasa
print (sheet["C1"].value)

#iteracio
for row in sheet.iter_rows(max_row=5):
    for cell in row:
        print(cell.value, end=" ")
    print()

#sor es oszlopok szamanak kiolvasasa
print ("Sorok szama: " + str(sheet.max_row))
print ("Oszlopok szama: " + str(sheet.max_column))

#oszlopnevek kiolvasasa
column_names = []
for column in sheet.iter_cols(1, sheet.max_column):
    column_names.append(column[0].value)

print (column_names)

#szotar keszites...



